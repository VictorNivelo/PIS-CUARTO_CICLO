from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from docx import Document
import PyPDF2
import pdfplumber
import openpyxl
from datetime import datetime
from django.db.models import Q
from PIS.models import (
    Genero,
    TipoDNI,
    Universidad,
    UsuarioPersonalizado,
    Facultad,
    Carrera,
    Ciclo,
    Materia,
    Genero,
)
from .forms import (
    GeneroForm,
    InformeCarreraForm,
    InformeCicloForm,
    InformeMateriaForm,
    RecuperarContraseniaForm,
    RegistrarUsuarioForm,
    UniversidadForm,
    FacultadForm,
    CarreraForm,
    CicloForm,
    MateriaForm,
    TipoDNIForm,
)


def PaginaPrincipal(request):
    return render(request, "Index.html")


@login_required
def PaginaAdministrador(request):
    return render(request, "PaginaAdministrador.html")


@login_required
def PaginaDocente(request):
    return render(request, "PaginaDocente.html")


@login_required
def PaginaSecretaria(request):
    return render(request, "PaginaSecretaria.html")


# Informacion del programa


def Galeria(request):
    return render(request, "Galeria.html")


def Informacion1(request):
    return render(request, "Informacion1.html")


def Informacion2(request):
    return render(request, "Informacion2.html")


def Informacion3(request):
    return render(request, "Informacion3.html")


def Informacion4(request):
    return render(request, "Informacion4.html")


# Informes


def InformeCiclo(request):
    if request.method == "POST":
        form = InformeCicloForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Informe de ciclo guardado exitosamente.")
            return redirect("Index")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = InformeCicloForm()
    return render(request, "InformeCiclo.html", {"form": form})


def InformeCarrera(request):
    if request.method == "POST":
        form = InformeCarreraForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Informe de carrera guardado exitosamente.")
            return redirect("Index")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = InformeCarreraForm()
    return render(request, "InformeCarrera.html", {"form": form})


def InformeMateria(request):
    if request.method == "POST":
        form = InformeMateriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request, "Informe de deserción estudiantil guardado exitosamente."
            )
            return redirect("Index")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = InformeMateriaForm()
    return render(request, "InformeMateria.html", {"form": form})


# Funciones del sistema


def Reporte(request):
    return render(request, "ReporteGenerado.html")


def Prediccion(request):
    return render(request, "Prediccion.html")


def PrediccionPresente(request):
    return render(request, "PrediccionPresente.html")


def Grafico(request):
    return render(request, "Grafico.html")


# Gestion de clases


def GestionUniversidad(request):
    return render(request, "GestionUniversidad.html")


def GestionFacultad(request):
    return render(request, "GestionFacultad.html")


def GestionCarrera(request):
    return render(request, "GestionCarrera.html")


def GestionCiclo(request):
    return render(request, "GestionCiclo.html")


def GestionMateria(request):
    return render(request, "GestionMateria.html")


# funciones de autenticacion


def RegistrarUsuario(request):
    if request.method == "POST":
        form = RegistrarUsuarioForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)

            if UsuarioPersonalizado.objects.count() == 0:
                user.is_superuser = True
                user.is_staff = True
                user.rol = "Personal Administrativo"
            else:
                user.rol = "Docente"

            user.set_password(form.cleaned_data["password1"])
            user.save()
            messages.success(request, "Usuario registrado exitosamente.")
            return redirect("Iniciar_Sesion")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == "username":
                        messages.error(
                            request, f"Error en el nombre de usuario: {error}"
                        )
                    elif field == "password1":
                        messages.error(request, f"Contraseña no segura: {error}")
                    elif field == "password2":
                        messages.error(
                            request, f"Las contraseñas no coinciden: {error}"
                        )
                    else:
                        messages.error(request, f"Error en el campo {field}: {error}")
    else:
        form = RegistrarUsuarioForm()

    return render(request, "RU-CrearUsuario.html", {"form": form})


def GestionUsuario(request):
    query = request.GET.get("search_query", "")
    filter_rol = request.GET.get("rol", "")
    filter_genero = request.GET.get("genero", "")
    filter_tipo_dni = request.GET.get("tipo_dni", "")

    usuarios = UsuarioPersonalizado.objects.all()
    generos = Genero.objects.all()
    tipos_dni = TipoDNI.objects.all()

    if query:
        usuarios = usuarios.filter(
            Q(username__icontains=query) | Q(dni__icontains=query)
        )
    if filter_rol:
        usuarios = usuarios.filter(rol=filter_rol)
    if filter_genero:
        usuarios = usuarios.filter(genero_id=filter_genero)
    if filter_tipo_dni:
        usuarios = usuarios.filter(tipo_dni=filter_tipo_dni)

    if request.method == "POST":
        if "modify" in request.POST:
            user_id = request.POST.get("user_id")
            usuario = UsuarioPersonalizado.objects.get(id=user_id)
            usuario.username = request.POST.get("username")
            usuario.first_name = request.POST.get("first_name")
            usuario.last_name = request.POST.get("last_name")
            usuario.rol = request.POST.get("rol")
            usuario.genero_id = request.POST.get("genero")
            usuario.tipo_dni_id = request.POST.get("tipo_dni")
            usuario.dni = request.POST.get("dni")
            fecha_nacimiento_str = request.POST.get("fecha_nacimiento")
            try:
                fecha_nacimiento = datetime.strptime(
                    fecha_nacimiento_str, "%d/%m/%Y"
                ).date()
                usuario.fecha_nacimiento = fecha_nacimiento
            except ValueError:
                messages.error(
                    request,
                    "Formato de fecha de nacimiento inválido. Utiliza el formato dd/mm/aaaa.",
                )
                return redirect("Gestion_Usuario")
            usuario.telefono = request.POST.get("telefono")
            usuario.save()
            messages.success(
                request, "Información del usuario actualizada correctamente."
            )
        elif "delete" in request.POST:
            user_id = request.POST.get("user_id")
            usuario = UsuarioPersonalizado.objects.get(id=user_id)
            usuario.delete()
            messages.success(request, "Usuario eliminado correctamente.")
        return redirect("Gestion_Usuario")

    return render(
        request,
        "GestionUsuario.html",
        {
            "usuarios": usuarios,
            "query": query,
            "filter_rol": filter_rol,
            "filter_genero": filter_genero,
            "filter_tipo_dni": filter_tipo_dni,
            "generos": generos,
            "tipos_dni": tipos_dni,
        },
    )


def IniciarSesion(request):
    if request.method == "POST":
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data.get("username")
            password = form.cleaned_data.get("password")
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                if user.rol == "Personal Administrativo":
                    return redirect("Pagina_Administrador")
                elif user.rol == "Docente":
                    return redirect("Pagina_Docente")
                elif user.rol == "Secretaria":
                    return redirect("Pagina_Secretaria")
                else:
                    messages.error(request, "Rol de usuario no reconocido.")
                # return redirect("Pagina_Usuario")
    else:
        form = AuthenticationForm()

    return render(request, "IniciarSesion.html", {"form": form})


@login_required
def CerrarSesion(request):
    logout(request)
    return redirect("Iniciar_Sesion")


def RecuperarContrasenia(request):
    if request.method == "POST":
        form = RecuperarContraseniaForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            messages.success(
                request,
                "Se ha enviado un enlace de recuperación a su correo electrónico.",
            )
            return redirect("Iniciar_Sesion")
        else:
            messages.error(request, "Error al enviar el enlace de recuperación.")
    else:
        form = RecuperarContraseniaForm()
    return render(request, "RecuperarContrasenia.html", {"form": form})


def RegistrarTipoDNI(request):
    if request.method == "POST":
        form = TipoDNIForm(request.POST)
        if form.is_valid():
            tipo_dni = form.save(commit=False)
            tipo_dni.save()
            messages.success(request, "Tipo de DNI registrado exitosamente.")
            return redirect("Gestion_TipoDNI")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = TipoDNIForm()
    return render(request, "RTD-CrearTipoDNI.html", {"form": form})


def GestionTipoDNI(request):
    query = request.GET.get("search_query", "")
    tipos_dni = TipoDNI.objects.all()

    if query:
        tipos_dni = tipos_dni.filter(Q(nombre_tipo_dni__icontains=query))

    if request.method == "POST":
        if "modify" in request.POST:
            tipo_dni_id = request.POST.get("tipo_dni_id")
            tipo_dni = TipoDNI.objects.get(id=tipo_dni_id)
            tipo_dni.nombre_tipo_dni = request.POST.get("nombre_tipo_dni")
            tipo_dni.descripcion_tipo_dni = request.POST.get("descripcion_tipo_dni")
            tipo_dni.save()
            messages.success(request, "Tipo de DNI actualizado exitosamente.")
        elif "delete" in request.POST:
            tipo_dni_id = request.POST.get("tipo_dni_id")
            tipo_dni = TipoDNI.objects.get(id=tipo_dni_id)
            tipo_dni.delete()
            messages.success(request, "Tipo de DNI eliminado exitosamente.")
        return redirect("Gestion_TipoDNI")

    return render(
        request, "GestionTipoDNI.html", {"tipos_dni": tipos_dni, "query": query}
    )


def RegistrarGenero(request):
    if request.method == "POST":
        form = GeneroForm(request.POST)
        if form.is_valid():
            genero = form.save(commit=False)
            genero.save()
            messages.success(request, "Género registrado exitosamente.")
            return redirect("Gestion_Genero")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = GeneroForm()
        # messages.success(request, "Género registrado exitosamente.")
        # return redirect("Gestion_Genero")
    return render(request, "RG-CrearGenero.html", {"form": form})


def GestionGenero(request):
    query = request.GET.get("search_query", "")
    generos = Genero.objects.all()

    if query:
        generos = generos.filter(Q(nombre_genero__icontains=query))

    if request.method == "POST":
        if "modify" in request.POST:
            genero_id = request.POST.get("genero_id")
            genero = Genero.objects.get(id=genero_id)
            genero.nombre_genero = request.POST.get("nombre_genero")
            genero.descripcion_genero = request.POST.get("descripcion_genero")
            genero.save()
            messages.success(request, "Género actualizado exitosamente.")
        elif "delete" in request.POST:
            genero_id = request.POST.get("genero_id")
            genero = Genero.objects.get(id=genero_id)
            genero.delete()
            messages.success(request, "Género eliminado exitosamente.")
        return redirect("Gestion_Genero")

    return render(request, "GestionGenero.html", {"generos": generos, "query": query})


def RegistrarUniversidad(request):
    if request.method == "POST":
        form = UniversidadForm(request.POST)
        if form.is_valid():
            universidad = form.save(commit=False)
            universidad.save()
            messages.success(request, "Universidad registrada exitosamente.")
            return redirect("Gestion_Universidad")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = UniversidadForm()

    return render(request, "RU-CrearUniversidad.html", {"form": form})


def GestionUniversidad(request):
    query = request.GET.get("search_query", "")

    universidades = Universidad.objects.all()

    if query:
        universidades = universidades.filter(
            Q(nombre_universidad__icontains=query)
            | Q(correo_universidad__icontains=query)
        )

    if request.method == "POST":
        if "modify" in request.POST:
            universidad_id = request.POST.get("universidad_id")
            universidad = Universidad.objects.get(id=universidad_id)
            universidad.nombre_universidad = request.POST.get("nombre_universidad")
            universidad.direccion_universidad = request.POST.get(
                "direccion_universidad"
            )
            universidad.telefono_universidad = request.POST.get("telefono_universidad")
            universidad.correo_universidad = request.POST.get("correo_universidad")
            fecha_fundacion_str = request.POST.get("fecha_fundacion")
            try:
                fecha_fundacion = datetime.strptime(
                    fecha_fundacion_str, "%d/%m/%Y"
                ).date()
                universidad.fecha_fundacion = fecha_fundacion
            except ValueError:
                messages.error(
                    request,
                    "Formato de fecha de nacimiento inválido. Utiliza el formato dd/mm/aaaa.",
                )
            universidad.save()
            messages.success(request, "Universidad actualizada exitosamente.")
        elif "delete" in request.POST:
            universidad_id = request.POST.get("universidad_id")
            universidad = Universidad.objects.get(id=universidad_id)
            universidad.delete()
            messages.success(request, "Universidad eliminada exitosamente.")
        return redirect("Gestion_Universidad")

    return render(
        request,
        "GestionUniversidad.html",
        {
            "universidades": universidades,
            "query": query,
        },
    )


def RegistrarFacultad(request):
    if request.method == "POST":
        form = FacultadForm(request.POST)
        if form.is_valid():
            facultad = form.save(commit=False)
            facultad.save()
            messages.success(request, "Facultad registrada exitosamente.")
            return redirect("Gestion_Facultad")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = FacultadForm()

    return render(request, "RF-CrearFacultad.html", {"form": form})


def GestionFacultad(request):
    query = request.GET.get("search_query", "")
    facultades = Facultad.objects.all()
    universidades = Universidad.objects.all()

    if query:
        facultades = facultades.filter(Q(nombre_facultad__icontains=query))

    if request.method == "POST":
        if "modify" in request.POST:
            facultad_id = request.POST.get("facultad_id")
            facultad = Facultad.objects.get(id=facultad_id)
            facultad.nombre_facultad = request.POST.get("nombre_facultad")

            fecha_fundacion_str = request.POST.get("fecha_fundacion")
            if fecha_fundacion_str:
                try:
                    fecha_fundacion = datetime.strptime(
                        fecha_fundacion_str, "%d/%m/%Y"
                    ).date()
                    facultad.fecha_fundacion = fecha_fundacion
                except ValueError:
                    messages.error(
                        request,
                        "Formato de fecha de fundación inválido. Utiliza el formato dd/mm/aaaa.",
                    )
                    return redirect("Gestion_Facultad")

            universidad_id = request.POST.get("universidad")
            universidad = Universidad.objects.get(id=universidad_id)
            facultad.universidad = universidad
            facultad.save()
            messages.success(request, "Facultad actualizada exitosamente.")
        elif "delete" in request.POST:
            facultad_id = request.POST.get("facultad_id")
            facultad = Facultad.objects.get(id=facultad_id)
            facultad.delete()
            messages.success(request, "Facultad eliminada exitosamente.")
        return redirect("Gestion_Facultad")

    return render(
        request,
        "GestionFacultad.html",
        {
            "facultades": facultades,
            "universidades": universidades,
            "query": query,
        },
    )


def RegistrarCarrera(request):
    if request.method == "POST":
        form = CarreraForm(request.POST)
        if form.is_valid():
            carrera = form.save(commit=False)
            carrera.save()
            messages.success(request, "Carrera registrada exitosamente.")
            return redirect("Gestion_Carrera")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = CarreraForm()

    return render(request, "RC-CrearCarrera.html", {"form": form})


def GestionCarrera(request):
    query = request.GET.get("search_query", "")
    carreras = Carrera.objects.all()
    facultades = Facultad.objects.all()

    if query:
        carreras = carreras.filter(Q(nombre_carrera__icontains=query))

    if request.method == "POST":
        if "modify" in request.POST:
            carrera_id = request.POST.get("carrera_id")
            carrera = Carrera.objects.get(id=carrera_id)
            carrera.nombre_carrera = request.POST.get("nombre_carrera")
            carrera.duracion = request.POST.get("duracion")
            facultad_id = request.POST.get("facultad")
            facultad = Facultad.objects.get(id=facultad_id)
            carrera.facultad = facultad
            carrera.save()
            messages.success(request, "Carrera actualizada exitosamente.")
        elif "delete" in request.POST:
            carrera_id = request.POST.get("carrera_id")
            carrera = Carrera.objects.get(id=carrera_id)
            carrera.delete()
            messages.success(request, "Carrera eliminada exitosamente.")
        return redirect("Gestion_Carrera")

    return render(
        request,
        "GestionCarrera.html",
        {
            "carreras": carreras,
            "facultades": facultades,
            "query": query,
        },
    )


def RegistrarCiclo(request):
    if request.method == "POST":
        form = CicloForm(request.POST)
        if form.is_valid():
            ciclo = form.save(commit=False)
            ciclo.save()
            messages.success(request, "Ciclo registrado exitosamente.")
            return redirect("Gestion_Ciclo")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = CicloForm()

    return render(request, "RC-CrearCiclo.html", {"form": form})


def GestionCiclo(request):
    query = request.GET.get("search_query", "")
    ciclos = Ciclo.objects.all()
    carreras = Carrera.objects.all()

    if query:
        ciclos = ciclos.filter(Q(nombre_ciclo__icontains=query))

    if request.method == "POST":
        if "modify" in request.POST:
            ciclo_id = request.POST.get("ciclo_id")
            ciclo = Ciclo.objects.get(id=ciclo_id)
            ciclo.nombre_ciclo = request.POST.get("nombre_ciclo")
            fecha_inicio_str = request.POST.get("fecha_inicio")
            if fecha_inicio_str:
                try:
                    fecha_inicio = datetime.strptime(
                        fecha_inicio_str, "%d/%m/%Y"
                    ).date()
                    ciclo.fecha_inicio = fecha_inicio
                except ValueError:
                    messages.error(
                        request,
                        "Formato de fecha de inicio inválido. Utiliza el formato dd/mm/aaaa.",
                    )
                    return redirect("Gestion_Ciclo")
            fecha_fin_str = request.POST.get("fecha_fin")
            if fecha_fin_str:
                try:
                    fecha_fin = datetime.strptime(fecha_fin_str, "%d/%m/%Y").date()
                    ciclo.fecha_fin = fecha_fin
                except ValueError:
                    messages.error(
                        request,
                        "Formato de fecha de fin inválido. Utiliza el formato dd/mm/aaaa.",
                    )
                    return redirect("Gestion_Ciclo")
            carrera_id = request.POST.get("carrera")
            carrera = Carrera.objects.get(id=carrera_id)
            ciclo.carrera = carrera
            ciclo.save()
            messages.success(request, "Ciclo actualizado exitosamente.")
        elif "delete" in request.POST:
            ciclo_id = request.POST.get("ciclo_id")
            ciclo = Ciclo.objects.get(id=ciclo_id)
            ciclo.delete()
            messages.success(request, "Ciclo eliminado exitosamente.")
        return redirect("Gestion_Ciclo")

    return render(
        request,
        "GestionCiclo.html",
        {
            "ciclos": ciclos,
            "carreras": carreras,
            "query": query,
        },
    )


def RegistrarMateria(request):
    if request.method == "POST":
        form = MateriaForm(request.POST)
        if form.is_valid():
            materia = form.save(commit=False)
            materia.save()
            messages.success(request, "Materia registrada exitosamente.")
            return redirect("Gestion_Materia")
        else:
            messages.error(request, "Por favor, corrija los errores del formulario.")
    else:
        form = MateriaForm()

    return render(request, "RM-CrearMateria.html", {"form": form})


def GestionMateria(request):
    query = request.GET.get("search_query", "")
    materias = Materia.objects.all()
    ciclos = Ciclo.objects.all()

    if query:
        materias = materias.filter(Q(nombre_materia__icontains=query))

    if request.method == "POST":
        if "modify" in request.POST:
            materia_id = request.POST.get("materia_id")
            materia = Materia.objects.get(id=materia_id)
            materia.nombre_materia = request.POST.get("nombre_materia")
            materia.save()
            messages.success(request, "Materia actualizada exitosamente.")
        elif "delete" in request.POST:
            materia_id = request.POST.get("materia_id")
            materia = Materia.objects.get(id=materia_id)
            materia.delete()
            messages.success(request, "Materia eliminada exitosamente.")
        return redirect("Gestion_Materia")

    return render(
        request,
        "GestionMateria.html",
        {
            "materias": materias,
            "ciclos": ciclos,
            "query": query,
        },
    )


# Funcionalidad de subir archivo para su procesamiento


def Extraer_DOCS(file):
    document = Document(file)
    data = {}
    for para in document.paragraphs:
        text = para.text.strip()
        if "Materia:" in text:
            data["materia"] = text.replace("Materia:", "").strip()
        elif "Docente encargado:" in text:
            data["docente_encargado"] = text.replace("Docente encargado:", "").strip()
        elif "Numero de estudiante:" in text:
            data["numero_estudiantes"] = int(
                text.replace("Numero de estudiante:", "").strip()
            )
        elif "Aprobados:" in text:
            data["aprobados"] = int(text.replace("Aprobados:", "").strip())
        elif "Reprobados:" in text:
            data["reprobados"] = int(text.replace("Reprobados:", "").strip())
        elif "Desertores:" in text:
            data["desertores"] = int(text.replace("Desertores:", "").strip())
        elif "Retirados:" in text:
            data["retirados"] = int(text.replace("Retirados:", "").strip())
    return data


def Extraer_PDF(file):
    data = {}
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                for line in text.split("\n"):
                    line = line.strip()
                    if "Materia:" in line:
                        data["materia"] = line.replace("Materia:", "").strip()
                    elif "Docente encargado:" in line:
                        data["docente_encargado"] = line.replace(
                            "Docente encargado:", ""
                        ).strip()
                    elif "Numero de estudiante:" in line:
                        data["numero_estudiantes"] = int(
                            line.replace("Numero de estudiante:", "").strip()
                        )
                    elif "Aprobados:" in line:
                        data["aprobados"] = int(line.replace("Aprobados:", "").strip())
                    elif "Reprobados:" in line:
                        data["reprobados"] = int(
                            line.replace("Reprobados:", "").strip()
                        )
                    elif "Desertores:" in line:
                        data["desertores"] = int(
                            line.replace("Desertores:", "").strip()
                        )
                    elif "Retirados:" in line:
                        data["retirados"] = int(line.replace("Retirados:", "").strip())
    return data


def Extraer_XLSX(file):
    data = {}
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, max_col=7, values_only=True):
        for cell_value in row:
            if isinstance(cell_value, str):
                if "Materia:" in cell_value:
                    data["materia"] = cell_value.replace("Materia:", "").strip()
                elif "Docente encargado:" in cell_value:
                    data["docente_encargado"] = cell_value.replace("Docente encargado:", "").strip()
            elif isinstance(cell_value, (int, float)):
                if "Numero de estudiante:" in str(cell_value):
                    data["numero_estudiantes"] = int(str(cell_value).replace("Numero de estudiante:", "").strip())
                elif "Aprobados:" in str(cell_value):
                    data["aprobados"] = int(str(cell_value).replace("Aprobados:", "").strip())
                elif "Reprobados:" in str(cell_value):
                    data["reprobados"] = int(str(cell_value).replace("Reprobados:", "").strip())
                elif "Desertores:" in str(cell_value):
                    data["desertores"] = int(str(cell_value).replace("Desertores:", "").strip())
                elif "Retirados:" in str(cell_value):
                    data["retirados"] = int(str(cell_value).replace("Retirados:", "").strip())

    workbook.close()
    return data

# def Extraer_XLSX(file):
#     data = {}
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook.active

#     for row in sheet.iter_rows(values_only=True):
#         for cell_value in row:
#             if isinstance(cell_value, str):
#                 if "Materia:" in cell_value:
#                     data["materia"] = cell_value.replace("Materia:", "").strip()
#                 elif "Docente encargado:" in cell_value:
#                     data["docente_encargado"] = cell_value.replace(
#                         "Docente encargado:", ""
#                     ).strip()
#             elif isinstance(cell_value, (int, float)):
#                 if "Numero de estudiante:" in str(cell_value):
#                     data["numero_estudiantes"] = int(
#                         str(cell_value).replace("Numero de estudiante:", "").strip()
#                     )
#                 elif "Aprobados:" in str(cell_value):
#                     data["aprobados"] = int(
#                         str(cell_value).replace("Aprobados:", "").strip()
#                     )
#                 elif "Reprobados:" in str(cell_value):
#                     data["reprobados"] = int(
#                         str(cell_value).replace("Reprobados:", "").strip()
#                     )
#                 elif "Desertores:" in str(cell_value):
#                     data["desertores"] = int(
#                         str(cell_value).replace("Desertores:", "").strip()
#                     )
#                 elif "Retirados:" in str(cell_value):
#                     data["retirados"] = int(
#                         str(cell_value).replace("Retirados:", "").strip()
#                     )

#     return data


def CargarInforme(request):
    if request.method == "POST":
        file = request.FILES["document"]
        file_extension = file.name.split(".")[-1].lower()

        if file_extension == "docx":
            data = Extraer_DOCS(file)
        elif file_extension == "pdf":
            data = Extraer_PDF(file)
        elif file_extension == "xlsx":
            data = Extraer_XLSX(file)
        else:
            return HttpResponse("Formato de archivo no soportado.", status=400)

        form = InformeMateriaForm(initial=data)
        return render(request, "InformeMateria.html", {"form": form})

    return render(request, "CargarInforme.html")
